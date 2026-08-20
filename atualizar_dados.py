#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Converte a planilha "Leds Urbs" (aba LEDS e aba Controle Norberto)
em dados JSON usados pelo dashboard (dashboard/dados.js).

Uso:
    python atualizar_dados.py "C:\caminho\para\Leds Urbs.xlsx"

Sem argumento, procura automaticamente pelo arquivo .xlsx mais recente em
Downloads, Documents e na pasta do script.
"""
import argparse
import datetime
import glob
import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("Faltando dependencia. Instale com:  pip install openpyxl")
    sys.exit(1)

SECTION_VEICULANDO = "veiculando"
SECTION_RESERVADAS = "reservadas"

EMPTY = {"", "----", "-", None}

def fmt_date(v):
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    if isinstance(v, (int, float)):
        return None
    return None

def parse_val(v):
    """Devolve (texto_bruto, data_iso)."""
    if v is None:
        return None, None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return None, fmt_date(v)
    if isinstance(v, (int, float)):
        return None, None
    s = str(v).strip()
    if s in EMPTY:
        return None, None
    return s, None

def clean_name(s):
    if s is None:
        return None
    return s.strip()

def find_latest_xlsx(paths):
    candidates = []
    for p in paths:
        if p and os.path.exists(p):
            if os.path.isdir(p):
                candidates.extend(glob.glob(os.path.join(p, "Leds Urbs*.xlsx")))
            else:
                candidates.append(p)
    if not candidates:
        return None
    return max(candidates, key=lambda f: os.path.getmtime(f))

def find_sections(ws):
    """Linhas dos titulos 'Campanhas veiculando'/'Campanhas reservadas'."""
    veic, res = [], []
    for r, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for v in row:
            if isinstance(v, str):
                s = v.strip().lower()
                if s == "campanhas veiculando":
                    veic.append(r)
                elif s == "campanhas reservadas":
                    res.append(r)
    return sorted(set(veic)), sorted(set(res))

def find_groups(ws):
    """Detecta grupos de telas: blocos lado a lado (nome + labels
    inicio/Fim/programacao) que se repetem verticalmente. Cada grupo
    comeca na linha de nomes, seguida dos titulos 'Campanhas veiculando'
    e 'Campanhas reservadas' que delimitam as secoes de campanhas."""
    veic, res = find_sections(ws)
    if not veic:
        return []
    groups = []
    for V in veic:
        names_row = V - 1
        labels = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=V + 1, column=c).value
            if not isinstance(v, str):
                continue
            s = v.strip().lower()
            if s == "inicio":
                labels[c] = "inicio"
            elif s == "fim":
                labels[c] = "fim"
            elif s.startswith("programa"):
                labels[c] = "prog"

        name_cols = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=names_row, column=c).value
            if v is not None and str(v).strip():
                name_cols.append((c, str(v).strip()))
        name_cols.sort(key=lambda x: x[0])

        blocks = []
        for i, (name_col, nome) in enumerate(name_cols):
            end = name_cols[i + 1][0] if i + 1 < len(name_cols) else ws.max_column + 1
            ini = fim = prog = None
            for c in range(name_col + 1, end):
                lab = labels.get(c)
                if lab == "inicio" and ini is None:
                    ini = c
                elif lab == "fim" and fim is None:
                    fim = c
                elif lab == "prog" and prog is None:
                    prog = c
            if ini is None:
                ini = name_col + 1
            if fim is None:
                fim = name_col + 2
            if prog is None:
                prog = name_col + 3
            blocks.append({"nome": nome, "name": name_col,
                           "ini": ini, "fim": fim, "prog": prog})

        resV = next((r for r in res if r > V), None)
        nextV = next((v2 for v2 in veic if v2 > V and (resV is None or v2 > resV)), None)
        groups.append({
            "blocks": blocks,
            "veic_ini": V + 2,
            "veic_fim": (resV - 1) if resV is not None else ws.max_row,
            "res_ini": (resV + 1) if resV is not None else None,
            "res_fim": (nextV - 2) if nextV is not None else ws.max_row,
        })
    return groups

def parse_leds_sheet(ws):
    locais = []
    for g in find_groups(ws):
        for b in g["blocks"]:
            local = {"numero": len(locais) + 1, "nome": b["nome"], "campanhas": []}

            def ler_linha(r):
                nome, _ = parse_val(ws.cell(row=r, column=b["name"]).value)
                ini_txt, ini_date = parse_val(ws.cell(row=r, column=b["ini"]).value)
                _, fim = parse_val(ws.cell(row=r, column=b["fim"]).value)
                prog, _ = parse_val(ws.cell(row=r, column=b["prog"]).value)
                nome = clean_name(nome)
                if nome is None:
                    return None
                return {
                    "nome": nome,
                    "inicio": ini_date,
                    "fim": fim,
                    "prog": prog,
                    "ini_texto": ini_txt,
                    "indice": None,
                }

            for r in range(g["veic_ini"], g["veic_fim"] + 1):
                c = ler_linha(r)
                if c:
                    c["secao"] = SECTION_VEICULANDO
                    local["campanhas"].append(c)

            if g["res_ini"] is not None:
                for r in range(g["res_ini"], g["res_fim"] + 1):
                    c = ler_linha(r)
                    if c:
                        c["secao"] = SECTION_RESERVADAS
                        local["campanhas"].append(c)

            locais.append(local)
    return locais

def parse_controle_sheet(ws):
    rows = []
    for row in ws.iter_rows(values_only=True):
        if not row or all(v is None for v in row):
            continue
        data = row[0]
        campanhas = row[1]
        regiao = row[2]
        n_regioes = row[3]
        if isinstance(data, (datetime.datetime, datetime.date)):
            data_str = data.isoformat()[:10]
        else:
            data_str = None
        if campanhas is None:
            continue
        rows.append({
            "data": data_str,
            "campanhas": str(campanhas).strip() if campanhas is not None else None,
            "regiao": str(regiao).strip() if regiao is not None else None,
            "n_regioes": int(n_regioes) if isinstance(n_regioes, (int, float)) else n_regioes,
        })
    return rows

def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("arquivo", nargs="?", help="caminho do arquivo .xlsx")
    parser.add_argument("-o", "--saida", default=os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        "dashboard", "dados.js"), help="arquivo de saida (default: dashboard/dados.js)")
    args = parser.parse_args()

    caminho = args.arquivo or find_latest_xlsx([
        os.path.join(os.path.expanduser("~"), "Downloads"),
        os.path.join(os.path.expanduser("~"), "Documents"),
        os.path.dirname(os.path.abspath(__file__)),
    ])
    if not caminho:
        print("Nenhum arquivo .xlsx encontrado. Passe o caminho: python atualizar_dados.py \"arquivo.xlsx\"")
        sys.exit(1)

    print(f"Lendo: {caminho}")
    wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    if "LEDS" not in wb.sheetnames:
        print("Aba 'LEDS' nao encontrada na planilha.")
        sys.exit(1)

    locais = parse_leds_sheet(wb["LEDS"])
    controle = parse_controle_sheet(wb["Controle Norberto"]) if "Controle Norberto" in wb.sheetnames else []
    wb.close()

    dados = {
        "fonte": os.path.basename(caminho),
        "gerado_em": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        "locais": locais,
        "controle": controle,
    }

    js = "// Gerado automaticamente por atualizar_dados.py - nao editar na mao\n"
    js += "window.DADOS_LEDS = " + json.dumps(dados, ensure_ascii=False, indent=2) + ";\n"

    with open(args.saida, "w", encoding="utf-8") as f:
        f.write(js)

    total = sum(len(l["campanhas"]) for l in locais)
    print(f"OK! {len(locais)} locais, {total} campanhas, {len(controle)} registros de controle.")
    print(f"Gerado: {args.saida}")

if __name__ == "__main__":
    main()
